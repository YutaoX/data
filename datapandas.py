#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu May 30 13:51:49 2019

@author: yutao
"""
import openpyxl
#import numpy as np
import pandas as pd
import os,sys
import matplotlib.pyplot as plt
"""
future_code = 'RB0'
os.chdir( "data" )
wb = openpyxl.load_workbook(future_code+".xlsx")
mavg1 = df['40_MA_Open'] = pd. Series.rolling(df['H'].values, 40)
mavg2 = df['60_MA_Open'] = pd.stats.moments.rolling_mean(df['H'].values, 60)
 pd.Series(df['H'].values).rolling(window=40).mean()
"""
#df=pd.read_excel('test.xlsx', usecols=[15, 16] ,header=0,names=["Date","Mpoint"],skiprows=200, skipfooter = 2967)
future_code = 'P0'
#wb = openpyxl.load_workbook(future_code+"any.xlsx")
wb = future_code+"any.xlsx"
#wb = "test.xlsx"
#对行号的获得及处理，以便构建df时没有空数据
wb2 = openpyxl.load_workbook(future_code+"any.xlsx")
ws = wb2.active 
print(ws)
row_max =ws.max_row
#print(rown)
num = 1
while 1:
     cell = ws.cell(row=num, column=14).value
     if cell:
         num = num +1         
     else:
        print(cell)
        break
#print(num)
rown= row_max-num+1
###对行号的获得及处理结束
df=pd.read_excel(wb, usecols=[13, 14] ,names=["Date","Mpoint"],skipfooter = rown)
#print(rown)
print(df)
df = pd.DataFrame(df['Mpoint'].values, index=df['Date'].values, columns=['Mpoint'])
#mavg1 = df['10_MA_Open'] = pd.Series(df['H'].values).rolling(window=10).mean()
#mavg2 = df['20_MA_Open'] = pd.Series(df['H'].values).rolling(window=20).mean()
##mavg = df['30_MA_Open'] = pd.stats.moments.rolling(df['Mpoint'],center=False,window=20).mean()
df.plot()
plt.show()
#===================================
#draw mavg lines as follows
#mavg1.plot()
#mavg2.plot()
#plt.show()
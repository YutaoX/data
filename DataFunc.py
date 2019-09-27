#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu May 30 13:51:49 2019

@author: yutao
数据的获取与分析函数集
Get_future_data_day(future_code)
Data_ansys(future_code)
Drawline_M(future_code)
"""
import openpyxl
#import numpy as np
import pandas as pd
import os,sys
import requests
import matplotlib.pyplot as plt
#创建数据获得的函数
def Get_future_data_day(future_code):
 #   future_code = 'RB0'
    url_str = ('http://stock2.finance.sina.com.cn/futures/api/json.php/IndexService.getInnerFuturesDailyKLine?symbol=' +
future_code)
    r = requests.get(url_str)
    r_json = r.json()
    r_lists = list(r_json)
    wb = openpyxl.Workbook()
    sheet = wb.active
        #sheet.title = 'sr0'
    sheet.title = future_code
    i=1
    for r_list in r_lists:
        j=0
        for v in r_list:
            sheet.cell(row=i+1, column=j+1, value=v)
            j=j+1
                #    print('\n')
        i=i+1
    wb.save(future_code+".xlsx")
    return
#创建数据分析的函数
def Data_ansys(future_code):
    #future_code = 'RB0'
#    os.chdir( "data" )
    #wb = openpyxl.load_workbook('RB0.xlsx')
    wb = openpyxl.load_workbook(future_code+".xlsx")
    
    #print(wb.get_sheet_names())
    # 根据sheet名字获得sheet
    #a_sheet = wb.get_sheet_by_name('RB0')
    a_sheet = wb.get_sheet_by_name(future_code)
    # 获得sheet名
    #print(a_sheet.title)
    # 获得当前正在显示的sheet, 也可以用wb.get_active_sheet()
    sheet = wb.active 
    #print(sheet.max_row)
    #print(sheet.max_column)
    rown=sheet.max_row
    #rown=5
    #str222=sheet.cell(row= 3, column=2).value[0:4]
    #tt=float(str222)
    #print(str222)
    a=2
    i=2
    sc=0
    scj=0
    bc=0
    bcj=0
    sheet.cell(row= a, column=8).value=0
    sheet.cell(row= a, column=9).value=0
    sheet.cell(row= a, column=10).value=0
    sheet.cell(row= a, column=11).value="SB"
    sheet.cell(row= a, column=14).value=sheet.cell(row=2, column=1).value
    sheet.cell(row= a, column=15).value=sheet.cell(row=2, column=5).value
    a=a+1
    sheet.cell(row= a, column=14).value=sheet.cell(row=2, column=1).value
    sheet.cell(row= a, column=15).value=sheet.cell(row=2, column=5).value
    while i<rown:
        while i<rown:
            cell15=float(sheet.cell(row= a, column=15).value)
            cel15=float(sheet.cell(row= a-1, column=15).value)
            celll15=float(sheet.cell(row= i+1, column=5).value)
            rs1 = cell15 > cel15 and cell15 > celll15
            rs2 = cell15 < cel15 and cell15 < celll15
    #        print(cell15)
    #        print(cel15)
    #        print(celll15)
            if (rs1 or rs2 ):
                a = a + 1
                sheet.cell(row= a, column=14).value= sheet.cell(row= i, column=1).value
                sheet.cell(row= a, column=15).value = sheet.cell(row= i, column=5).value
                break
            i=i+1    
            sopen=float(sheet.cell(row= i, column=2).value)
            sclose=float(sheet.cell(row= i, column=5).value)
            sclosey=float(sheet.cell(row= i-1, column=5).value)
            smax=float(sheet.cell(row= i, column=3).value)
            smin=float(sheet.cell(row= i, column=4).value)
            sheet.cell(row= i, column=8).value=smax-smin
            sheet.cell(row= i, column=9).value=sopen-sclose
            sheet.cell(row= i, column=10).value=sclose-sclosey
            sheet.cell(row= a, column=14).value=sheet.cell(row=i, column=1).value
            sheet.cell(row= a, column=15).value=sheet.cell(row=i, column=5).value
            ssb=float(sheet.cell(row= i, column=10).value)
            if ssb==0:
                sbflag="SB"            
            elif ssb>0:
                sbflag="B"
                bc=bc+1
                bcj=bcj+ssb
            else:
                sbflag="S"
                sc=sc+1
                scj=scj+ssb            
            sheet.cell(row= i, column=11).value=sbflag 
    #        print(i)
    wb.save(future_code+"any.xlsx")
    print(future_code ,"datashow")    
    print("sell:",sc,scj,"buy:",bc,bcj)
    print("sell:",int(scj/sc),"buy:",int(bcj/bc))
    return
####################################################
def Drawline_M(future_code):
    wb = future_code+"any.xlsx"
    #wb = "test.xlsx"
    #对行号的获得及处理，以便构建df时没有空数据
    wb2 = openpyxl.load_workbook(future_code+"any.xlsx")
    ws = wb2.active 
    print(ws)
    row_max =ws.max_row
    #print(rown)
    num = 1
    while 1:
         cell = ws.cell(row=num, column=14).value
         if cell:
             num = num +1         
         else:
            print(cell)
            break
    #print(num)
    rown= row_max-num+1
    ###对行号的获得及处理结束
    df=pd.read_excel(wb, usecols=[13, 14] ,names=["Date","Mpoint"],skipfooter = rown)
    #print(rown)
    print(df)
    df = pd.DataFrame(df['Mpoint'].values, index=df['Date'].values, columns=['Mpoint'])
    #mavg1 = df['10_MA_Open'] = pd.Series(df['Mpoint'].values).rolling(window=10).mean()
    #mavg2 = df['20_MA_Open'] = pd.Series(df['Mpoint'].values).rolling(window=20).mean()
    ##mavg = df['30_MA_Open'] = pd.stats.moments.rolling(df['Mpoint'],center=False,window=20).mean()
    df.plot()
    plt.show()
    return
#######################################
def Datacode(future_code):
    code=future_code
    Get_future_data_day(code)
    Data_ansys(code)
    Drawline_M(code)
    return
